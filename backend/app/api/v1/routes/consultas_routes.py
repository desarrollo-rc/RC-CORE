# backend/app/api/v1/routes/consultas_routes.py
from flask import Blueprint, request, jsonify, send_file
from flask_jwt_extended import jwt_required, get_jwt_identity
from app.api.v1.schemas.consulta_schemas import ConsultaSchema, ConsultaUpdateSchema, EjecucionParametrosSchema
from app.api.v1.services.consulta_service import ConsultaService
from app.api.v1.utils.decorators import permission_required
from werkzeug.exceptions import NotFound, BadRequest, Forbidden
from io import BytesIO

consultas_bp = Blueprint('consultas_bp', __name__)

schema_single = ConsultaSchema()
schema_many = ConsultaSchema(many=True)
schema_update = ConsultaUpdateSchema()
schema_params = EjecucionParametrosSchema()

@consultas_bp.route('/', methods=['GET'])
@jwt_required()
@permission_required('consultas:ver')
def list_consultas():
    try:
        consultas = ConsultaService.list_consultas()
        return schema_many.dump(consultas), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@consultas_bp.route('/', methods=['POST'])
@jwt_required()
@permission_required('consultas:crear')
def create_consulta():
    try:
        data = schema_single.load(request.get_json())
        id_usuario = get_jwt_identity()
        consulta = ConsultaService.create_consulta(data, id_usuario)
        return schema_single.dump(consulta), 201
    except Exception as e:
        return jsonify({"error": str(e)}), 400

@consultas_bp.route('/<string:codigo_consulta>', methods=['GET'])
@jwt_required()
@permission_required('consultas:ver')
def get_consulta(codigo_consulta):
    try:
        consulta = ConsultaService.get_consulta_by_codigo(codigo_consulta)
        return schema_single.dump(consulta), 200
    except NotFound as e:
        return jsonify({"error": str(e)}), 404

@consultas_bp.route('/<string:codigo_consulta>', methods=['PUT'])
@jwt_required()
@permission_required('consultas:editar')
def update_consulta(codigo_consulta):
    try:
        data = schema_update.load(request.get_json())
        id_usuario = get_jwt_identity()
        consulta = ConsultaService.update_consulta(codigo_consulta, data, id_usuario)
        return schema_single.dump(consulta), 200
    except NotFound as e:
        return jsonify({"error": str(e)}), 404
    except Exception as e:
        import traceback
        print(f"Error al actualizar consulta: {str(e)}")
        print(traceback.format_exc())
        return jsonify({"error": str(e)}), 400

@consultas_bp.route('/<string:codigo_consulta>', methods=['DELETE'])
@jwt_required()
@permission_required('consultas:eliminar')
def delete_consulta(codigo_consulta):
    try:
        ConsultaService.delete_consulta(codigo_consulta)
        return jsonify({"message": "Consulta eliminada exitosamente"}), 200
    except NotFound as e:
        return jsonify({"error": str(e)}), 404
    except Exception as e:
        return jsonify({"error": str(e)}), 400

# --- NUEVOS ENDPOINTS ESTRATÉGICOS ---

@consultas_bp.route('/<string:codigo_consulta>/execute', methods=['POST'])
@jwt_required()
@permission_required('consultas:ejecutar')
def execute_consulta(codigo_consulta):
    try:
        id_usuario = get_jwt_identity()
        # Validamos el body, aunque esté vacío
        params_data = schema_params.load(request.get_json() or {})
        resultado = ConsultaService.execute_by_codigo(codigo_consulta, id_usuario, params_data.get('parametros'))
        return jsonify(resultado), 200
    except (NotFound, Forbidden, BadRequest) as e:
        return jsonify({"error": e.description}), e.code
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# --- PROCESO DE NEGOCIO DE ALTO NIVEL ---

@consultas_bp.route('/procesos/sincronizar-pedidos-b2b', methods=['POST'])
@jwt_required()
@permission_required('procesos:b2b') # Un permiso específico para esta acción
def sync_pedidos():
    try:
        id_usuario = get_jwt_identity()
        resultado = ConsultaService.sincronizar_estados_pedidos_b2b(id_usuario)
        return jsonify(resultado), 200
    except (BadRequest, Forbidden) as e:
        return jsonify({"error": e.description}), e.code
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@consultas_bp.route('/<string:codigo_consulta>/download-excel', methods=['POST'])
@jwt_required()
@permission_required('consultas:ejecutar')
def download_consulta_excel(codigo_consulta):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        
        id_usuario = get_jwt_identity()
        params_data = schema_params.load(request.get_json() or {})
        
        # Ejecutar la consulta sin límite
        resultado = ConsultaService.execute_by_codigo_for_excel(codigo_consulta, id_usuario, params_data.get('parametros'))
        
        if not resultado.get('data'):
            return jsonify({"error": "La consulta no arrojó resultados"}), 404
        
        # Crear el workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Resultados"
        
        # Obtener las cabeceras
        headers = list(resultado['data'][0].keys())
        
        # Estilo para las cabeceras
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        # Escribir cabeceras
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
        
        # Escribir datos
        for row_idx, row_data in enumerate(resultado['data'], start=2):
            for col_idx, header in enumerate(headers, start=1):
                ws.cell(row=row_idx, column=col_idx, value=row_data.get(header))
        
        # Ajustar ancho de columnas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Guardar en memoria
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'{codigo_consulta}_resultados.xlsx'
        )
        
    except (NotFound, Forbidden, BadRequest) as e:
        return jsonify({"error": e.description}), e.code
    except Exception as e:
        return jsonify({"error": str(e)}), 500